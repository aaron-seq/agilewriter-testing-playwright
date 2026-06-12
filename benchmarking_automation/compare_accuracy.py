"""
Accuracy Comparison Script
===========================
Source of truth: QA report (tests/ICF_docx/QA report_ICF_FULL_0804 - Copy - Copy (2).xlsx)
Compares against: Pipeline output (final_outputs/replacement_inventory.xlsx)

Shows all QA entries with their expected columns, then checks if the pipeline
produced matching data and highlights what's missing or incorrect.

Output: accuracy_report.xlsx
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from pathlib import Path
import re
from difflib import SequenceMatcher

QA_REPORT = Path(r"tests/ICF_docx/QA report_ICF_FULL_0804 - Copy - Copy (2).xlsx")
PIPELINE_OUTPUT = Path("final_outputs/replacement_inventory.xlsx")
OUTPUT_FILE = Path("final_outputs/accuracy_report.xlsx")

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
BOLD_FONT = Font(bold=True, size=11)
NORMAL_FONT = Font(size=10)
THIN_BORDER = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
PLACEHOLDER_RE = re.compile(r"<[^>]+>")


def normalize(text):
    if text is None:
        return ""
    return re.sub(r"\s+", " ", str(text).strip()).lower()


def normalize_type(t):
    return normalize(t).replace("_", "").replace("-", "").replace(" ", "")


def similarity(a, b):
    a, b = normalize(a), normalize(b)
    if not a and not b:
        return 1.0
    if not a or not b:
        return 0.0
    return round(SequenceMatcher(None, a, b).ratio(), 4)


# ────────────────────────────────────────────────────────────────
# 1. Extract ALL QA entries (rows with placeholders and their columns)
# ────────────────────────────────────────────────────────────────
def extract_qa_all():
    """
    Return list of dicts: {placeholder, type, expected_value, ai_text, ai_source, matching_accuracy, similarity_score, source_match, ...}
    Reads all columns from the QA sheet rows that have a placeholder.
    """
    wb = openpyxl.load_workbook(QA_REPORT, data_only=True)
    ws = wb["QA"]
    header_row = 2  # headers are in row 2

    headers = [c.value for c in ws[header_row]]
    # Map column names to indices
    col_map = {}
    for idx, h in enumerate(headers):
        if h:
            col_map[str(h).strip().lower()] = idx

    qa_all = []

    for row_idx in range(header_row + 1, ws.max_row + 1):
        row = [c.value for c in ws[row_idx]]
        ph = row[col_map.get("placeholder")] if col_map.get("placeholder") is not None and col_map["placeholder"] < len(row) else None

        # Skip rows without a valid placeholder
        if ph is None or not isinstance(ph, str) or not PLACEHOLDER_RE.search(ph):
            continue
        if ph.strip() in ("#VALUE!", ""):
            continue

        entry = {
            "qa_placeholder": ph.strip(),
            "qa_type": str(row[col_map.get("placeholder type")]).strip() if col_map.get("placeholder type") is not None and col_map["placeholder type"] < len(row) and row[col_map["placeholder type"]] else "",
            "qa_expected_value": str(row[col_map.get("expected value")]).strip() if col_map.get("expected value") is not None and col_map["expected value"] < len(row) and row[col_map["expected value"]] else "",
            "qa_source": str(row[col_map.get("source document")]).strip() if col_map.get("source document") is not None and col_map["source document"] < len(row) and row[col_map["source document"]] else "",
            "qa_writing_instruction": str(row[col_map.get("writing instruction")]).strip() if col_map.get("writing instruction") is not None and col_map["writing instruction"] < len(row) and row[col_map["writing instruction"]] else "",
            "qa_ai_text": str(row[col_map.get("ai replaced text")]).strip() if col_map.get("ai replaced text") is not None and col_map["ai replaced text"] < len(row) and row[col_map["ai replaced text"]] else "",
            "qa_ai_source": str(row[col_map.get("ai detected source ")]).strip() if col_map.get("ai detected source ") is not None and col_map["ai detected source "] < len(row) and row[col_map["ai detected source "]] else "",
            "qa_similarity_score": row[col_map.get("similarity score")] if col_map.get("similarity score") is not None and col_map["similarity score"] < len(row) else None,
        }

        # Clean up empty/None values
        for k in entry:
            if isinstance(entry[k], str) and entry[k].strip() in ("None", "0", ""):
                entry[k] = ""

        # Normalize type for matching
        entry["qa_type_norm"] = normalize_type(entry["qa_type"])
        entry["qa_ph_norm"] = normalize(entry["qa_placeholder"])

        qa_all.append(entry)

    wb.close()
    return qa_all


# ────────────────────────────────────────────────────────────────
# 2. Extract pipeline entries grouped by placeholder
# ────────────────────────────────────────────────────────────────
def extract_pipeline_index():
    """Build a lookup: normalized placeholder -> list of pipeline entries."""
    wb = openpyxl.load_workbook(PIPELINE_OUTPUT, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    ph_col = next(i for i, h in enumerate(headers) if h and "placeholder" in h.lower())
    type_col = next(i for i, h in enumerate(headers) if h and h == "Type")
    repl_col = next(i for i, h in enumerate(headers) if h and "replacement content" in h.lower())
    status_col = next(i for i, h in enumerate(headers) if h and "status" in h.lower())

    index = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        ph = row[ph_col]
        if ph is None:
            continue
        ph_norm = normalize(str(ph).strip())
        repl = str(row[repl_col]).strip() if repl_col is not None and row[repl_col] else ""
        if repl == "None":
            repl = ""
        entry = {
            "placeholder": str(ph).strip(),
            "type": str(row[type_col]).strip() if type_col is not None and row[type_col] else "",
            "replacement": repl,
            "status": str(row[status_col]).strip() if status_col is not None and row[status_col] else "",
        }
        if ph_norm not in index:
            index[ph_norm] = []
        index[ph_norm].append(entry)

    wb.close()
    return index


# ────────────────────────────────────────────────────────────────
# 3. Compare QA (source of truth) vs Pipeline output
# ────────────────────────────────────────────────────────────────
def compare(qa_all, pipeline_index):
    """
    For each QA entry, find matching pipeline entries.
    Report what exists in QA but is missing or different in pipeline.
    Also report pipeline entries not found in QA.
    """
    results = []
    matched_pipeline_keys = set()
    qa_with_pipeline = set()

    for qa in qa_all:
        ph_norm = qa["qa_ph_norm"]
        pipe_entries = pipeline_index.get(ph_norm, [])

        # Find best match among all pipeline entries for this placeholder
        matching_pipe = None
        best_score = 0.0
        for pe in pipe_entries:
            score = similarity(pe["replacement"], qa["qa_ai_text"])
            if score > best_score:
                best_score = score
                matching_pipe = pe
                if score == 1.0:
                    break

        if matching_pipe:
            qa_with_pipeline.add(ph_norm)
            matched_pipeline_keys.add((ph_norm, matching_pipe["replacement"]))

            # Type comparison
            pipe_type_norm = normalize_type(matching_pipe["type"])
            type_verdict = "YES" if pipe_type_norm == qa["qa_type_norm"] else "NO"

            # Content comparison
            if matching_pipe["replacement"] and qa["qa_ai_text"]:
                if best_score == 1.0:
                    content_verdict = "MATCH"
                elif best_score >= 0.8:
                    content_verdict = "PARTIAL"
                else:
                    content_verdict = "MISMATCH"
            elif not matching_pipe["replacement"] and not qa["qa_ai_text"]:
                content_verdict = "BOTH EMPTY"
            elif not matching_pipe["replacement"]:
                content_verdict = "MISSING IN PIPELINE"
            else:
                content_verdict = "MISSING IN QA"

            results.append({
                "qa_placeholder": qa["qa_placeholder"],
                "qa_type": qa["qa_type"],
                "qa_ai_text": qa["qa_ai_text"],
                "pipeline_placeholder": matching_pipe["placeholder"],
                "pipeline_type": matching_pipe["type"],
                "pipeline_replacement": matching_pipe["replacement"],
                "type_match": type_verdict,
                "content_match": content_verdict,
                "similarity": best_score,
                "pipeline_status": matching_pipe["status"],
                "qa_expected": qa["qa_expected_value"],
                "qa_similarity": qa["qa_similarity_score"],
                "section": "QA present in Pipeline",
            })
        else:
            # QA entry has no match in pipeline output at all
            results.append({
                "qa_placeholder": qa["qa_placeholder"],
                "qa_type": qa["qa_type"],
                "qa_ai_text": qa["qa_ai_text"],
                "pipeline_placeholder": "",
                "pipeline_type": "",
                "pipeline_replacement": "",
                "type_match": "N/A",
                "content_match": "NOT FOUND IN PIPELINE",
                "similarity": 0.0,
                "pipeline_status": "",
                "qa_expected": qa["qa_expected_value"],
                "qa_similarity": qa["qa_similarity_score"],
                "section": "MISSING from Pipeline",
            })

    # Find pipeline entries that are NOT in QA
    for ph_norm, pipe_entries in pipeline_index.items():
        if ph_norm not in qa_with_pipeline:
            for pe in pipe_entries:
                results.append({
                    "qa_placeholder": "",
                    "qa_type": "",
                    "qa_ai_text": "",
                    "pipeline_placeholder": pe["placeholder"],
                    "pipeline_type": pe["type"],
                    "pipeline_replacement": pe["replacement"],
                    "type_match": "N/A",
                    "content_match": "EXTRA IN PIPELINE",
                    "similarity": 0.0,
                    "pipeline_status": pe["status"],
                    "qa_expected": "",
                    "qa_similarity": None,
                    "section": "Extra - not in QA",
                })

    # Summary
    total_qa = len(qa_all)
    total_pipeline = sum(len(v) for v in pipeline_index.values())
    found = sum(1 for r in results if r["section"] == "QA present in Pipeline")
    missing = sum(1 for r in results if r["section"] == "MISSING from Pipeline")
    extra = sum(1 for r in results if r["section"] == "Extra - not in QA")

    type_compared = sum(1 for r in results if r["type_match"] in ("YES", "NO"))
    type_matched = sum(1 for r in results if r["type_match"] == "YES")

    content_compared = sum(1 for r in results if r["content_match"] in ("MATCH", "PARTIAL", "MISMATCH"))
    content_match_count = sum(1 for r in results if r["content_match"] == "MATCH")
    content_partial = sum(1 for r in results if r["content_match"] == "PARTIAL")
    content_mismatch = sum(1 for r in results if r["content_match"] == "MISMATCH")
    content_missing = sum(1 for r in results if r["content_match"] == "MISSING IN PIPELINE")

    summary = {
        "total_qa": total_qa,
        "total_pipeline": total_pipeline,
        "qa_found_in_pipeline": found,
        "qa_missing_from_pipeline": missing,
        "pipeline_extra_not_in_qa": extra,
        "placeholder_coverage": round(found / total_qa * 100, 2) if total_qa else 0,
        "type_compared": type_compared,
        "type_matched": type_matched,
        "type_accuracy": round(type_matched / type_compared * 100, 2) if type_compared else 0,
        "content_compared": content_compared,
        "content_match": content_match_count,
        "content_partial": content_partial,
        "content_mismatch": content_mismatch,
        "content_missing_in_pipeline": content_missing,
        "content_match_rate": round(content_match_count / content_compared * 100, 2) if content_compared else 0,
    }

    return results, summary


# ────────────────────────────────────────────────────────────────
# 4. Write to Excel
# ────────────────────────────────────────────────────────────────
def write_output(results, summary):
    wb = openpyxl.Workbook()

    # ---- Sheet 1: Full Comparison (QA as source of truth) ----
    ws = wb.active
    ws.title = "QA vs Pipeline Comparison"

    headers = [
        "QA Placeholder",
        "QA Placeholder Type",
        "QA AI Replaced Text",
        "QA Expected Value",
        "QA Similarity Score",
        "Pipeline Placeholder",
        "Pipeline Type",
        "Pipeline Replacement",
        "Type Match",
        "Content Match",
        "Similarity",
        "Pipeline Status",
        "Section",
    ]

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN_BORDER

    for ri, r in enumerate(results, 2):
        data = [
            r["qa_placeholder"],
            r["qa_type"],
            r["qa_ai_text"],
            r["qa_expected"],
            r["qa_similarity"],
            r["pipeline_placeholder"],
            r["pipeline_type"],
            r["pipeline_replacement"],
            r["type_match"],
            r["content_match"],
            r["similarity"],
            r["pipeline_status"],
            r["section"],
        ]
        for ci, val in enumerate(data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Color code
            if ci == 9:  # Type Match
                cell.fill = GREEN_FILL if val == "YES" else (RED_FILL if val == "NO" else PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"))
            if ci == 10:  # Content Match
                if val == "MATCH":
                    cell.fill = GREEN_FILL
                elif val == "PARTIAL":
                    cell.fill = YELLOW_FILL
                elif val in ("MISMATCH", "MISSING IN PIPELINE", "NOT FOUND IN PIPELINE"):
                    cell.fill = RED_FILL
            if ci == 13:  # Section
                if "MISSING" in str(val):
                    cell.fill = RED_FILL
                elif "Extra" in str(val):
                    cell.fill = YELLOW_FILL

    ws.freeze_panes = "A2"
    widths = [35, 16, 50, 50, 14, 35, 16, 50, 12, 20, 12, 14, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "AA"].width = w

    # ---- Sheet 2: Missing from Pipeline (QA entries pipeline didn't capture) ----
    ws2 = wb.create_sheet("Missing in Pipeline Output")
    for ci, h in enumerate(["QA Placeholder", "QA Type", "QA AI Replaced Text", "QA Expected Value"], 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.border = THIN_BORDER

    ri = 2
    for r in results:
        if r["section"] == "MISSING from Pipeline":
            for ci, val in enumerate([r["qa_placeholder"], r["qa_type"], r["qa_ai_text"], r["qa_expected"]], 1):
                cell = ws2.cell(row=ri, column=ci, value=val)
                cell.font = NORMAL_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            ri += 1

    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 60
    ws2.column_dimensions["D"].width = 60

    # ---- Sheet 3: Extra in Pipeline (not in QA) ----
    ws3 = wb.create_sheet("Extra in Pipeline (not in QA)")
    for ci, h in enumerate(["Pipeline Placeholder", "Pipeline Type", "Pipeline Replacement", "Pipeline Status"], 1):
        c = ws3.cell(row=1, column=ci, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.border = THIN_BORDER

    ri = 2
    for r in results:
        if r["section"] == "Extra - not in QA":
            for ci, val in enumerate([r["pipeline_placeholder"], r["pipeline_type"], r["pipeline_replacement"], r["pipeline_status"]], 1):
                cell = ws3.cell(row=ri, column=ci, value=val)
                cell.font = NORMAL_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            ri += 1

    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 16
    ws3.column_dimensions["C"].width = 60
    ws3.column_dimensions["D"].width = 14

    # ---- Sheet 4: Summary ----
    ws4 = wb.create_sheet("Summary")
    items = [
        ("METRIC", "VALUE"),
        ("", ""),
        ("QA REPORT (Source of Truth)", ""),
        ("Total QA Entries (with placeholders)", summary["total_qa"]),
        ("", ""),
        ("PIPELINE COVERAGE", ""),
        ("QA entries found in Pipeline output", summary["qa_found_in_pipeline"]),
        ("QA entries MISSING from Pipeline output", summary["qa_missing_from_pipeline"]),
        ("Pipeline entries not in QA (extras)", summary["pipeline_extra_not_in_qa"]),
        ("Placeholder Coverage Rate", f"{summary['placeholder_coverage']}%"),
        ("", ""),
        ("TYPE MATCH ACCURACY (Pipeline Type vs QA Type)", ""),
        ("Type comparisons done", summary["type_compared"]),
        ("Type matches", summary["type_matched"]),
        ("Type accuracy", f"{summary['type_accuracy']}%"),
        ("", ""),
        ("CONTENT MATCH (Pipeline Replacement vs QA AI Replaced Text)", ""),
        ("Content comparisons done (both have data)", summary["content_compared"]),
        ("Exact MATCH", summary["content_match"]),
        ("Partial (>=80%)", summary["content_partial"]),
        ("MISMATCH", summary["content_mismatch"]),
        ("QA has content, Pipeline empty", summary["content_missing_in_pipeline"]),
        ("Exact match rate", f"{summary['content_match_rate']}%"),
        ("", ""),
        ("SUMMARY", ""),
        (f"Out of {summary['total_qa']} QA entries:", ""),
        (f"  - {summary['qa_found_in_pipeline']} found in pipeline output", ""),
        (f"  - {summary['qa_missing_from_pipeline']} MISSING from pipeline output", ""),
        (f"  - {summary['pipeline_extra_not_in_qa']} extra in pipeline not in QA", ""),
    ]

    for ri, (label, val) in enumerate(items, 1):
        cl = ws4.cell(row=ri, column=1, value=label)
        cv = ws4.cell(row=ri, column=2, value=val)
        cl.border = THIN_BORDER
        cv.border = THIN_BORDER
        if ri == 1:
            cl.fill = HEADER_FILL; cl.font = HEADER_FONT
            cv.fill = HEADER_FILL; cv.font = HEADER_FONT
        if isinstance(val, str) and "%" in val:
            try:
                nv = float(val.replace("%", ""))
                cv.fill = GREEN_FILL if nv >= 80 else (YELLOW_FILL if nv >= 50 else RED_FILL)
                cv.font = BOLD_FONT
            except ValueError:
                pass

    ws4.column_dimensions["A"].width = 55
    ws4.column_dimensions["B"].width = 25

    wb.save(OUTPUT_FILE)
    print(f"\nSaved to: {OUTPUT_FILE}")

    # Print summary
    print("\n" + "=" * 65)
    print("ACCURACY RESULTS (QA = Source of Truth)")
    print("=" * 65)
    print(f"QA entries (source of truth): {summary['total_qa']}")
    print(f"")
    print(f"COVERAGE:")
    print(f"  Found in pipeline:    {summary['qa_found_in_pipeline']}")
    print(f"  MISSING from pipeline: {summary['qa_missing_from_pipeline']}")
    print(f"  Extra in pipeline:     {summary['pipeline_extra_not_in_qa']}")
    print(f"  Coverage rate:         {summary['placeholder_coverage']}%")
    print(f"")
    print(f"TYPE MATCH:")
    print(f"  Accuracy: {summary['type_matched']}/{summary['type_compared']} ({summary['type_accuracy']}%)")
    print(f"")
    print(f"CONTENT MATCH (Replacement vs AI Text):")
    print(f"  Compared (both have content): {summary['content_compared']}")
    print(f"  Exact MATCH:                 {summary['content_match']}")
    print(f"  Partial (>=80%):             {summary['content_partial']}")
    print(f"  MISMATCH:                    {summary['content_mismatch']}")
    print(f"  QA has content, Pipeline empty: {summary['content_missing_in_pipeline']}")
    print(f"  Exact match rate:            {summary['content_match_rate']}%")
    print("=" * 65)


# ────────────────────────────────────────────────────────────────
# Main
# ────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("Reading QA report (source of truth)...")
    qa_all = extract_qa_all()
    print(f"  Extracted {len(qa_all)} entries from QA sheet")

    print("Reading pipeline output...")
    pipeline_index = extract_pipeline_index()
    total_pipe = sum(len(v) for v in pipeline_index.values())
    print(f"  Found {total_pipe} entries in pipeline output ({len(pipeline_index)} unique placeholders)")

    print("Comparing...")
    results, summary = compare(qa_all, pipeline_index)

    print("Writing report...")
    write_output(results, summary)