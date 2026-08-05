"""
Write the extracted placeholders as a QA workbook.

The shape matches the raw QA files in raw_qa_files/ so that
tests/helpers/accuracy-scorer.ts can read the output unchanged. That parser
reads the "QA" sheet from row 3 onwards and takes:

    col 0  Placeholder          col 4  Source Document
    col 2  Placeholder Type     col 5  Writing instruction
    col 6  AI Replaced Text

Only the columns that describe the placeholder are filled. Expected Value and
AI Replaced Text are left empty on purpose - QA fills those in.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .classify import KEY_VALUE, LIST, PARAGRAPH, TABLE, UNKNOWN, classify
from .extractor import Placeholder

# Column order of the per-type sheets in the real QA workbooks. "Column1" is
# blank in the originals but the positions after it are load-bearing, so it
# stays.
HEADERS = [
    "Placeholder",
    "Column1",
    "Placeholder Type",
    "Expected Value",
    "Source Document",
    "Writing instruction",
    "AI Replaced Text",
    "Ai detected Source ",
    "Matching Accuracy",
    "Similarity Score",
    "Source Match",
]

# Sheets the real workbooks carry, in order. Inline and Multi Tables exist but
# are usually empty; they are created so the shape matches.
TYPE_SHEETS = ["Inline", "Paragraph", "KeyValue", "Table", "Lists", "Multi Tables"]

_SHEET_FOR_TYPE = {
    PARAGRAPH: "Paragraph",
    KEY_VALUE: "KeyValue",
    TABLE: "Table",
    LIST: "Lists",
    UNKNOWN: "Paragraph",
}

_HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
_HEADER_FONT = Font(bold=True)


def build(placeholders: list[Placeholder], template_name: str) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)

    rows = [_row_for(p) for p in placeholders]

    _write_summary(workbook, placeholders, template_name)
    _write_qa_sheet(workbook, rows)

    for sheet_name in TYPE_SHEETS:
        matching = [r for r in rows if _SHEET_FOR_TYPE.get(r["type"]) == sheet_name]
        _write_type_sheet(workbook, sheet_name, matching)

    return workbook


def _row_for(placeholder: Placeholder) -> dict:
    return {
        "name": f"<{placeholder.name}>",
        "type": classify(placeholder.name, placeholder.structure),
        # The surrounding sentence is the best available writing instruction:
        # it is what a human reads to work out what belongs here.
        "instruction": placeholder.context,
        "heading": placeholder.heading,
        "count": placeholder.count,
        "structure": placeholder.structure,
    }


def _write_summary(workbook: Workbook, placeholders: list[Placeholder], template_name: str) -> None:
    sheet = workbook.create_sheet("Summary")
    counts: dict[str, int] = {}
    for p in placeholders:
        key = classify(p.name, p.structure)
        counts[key] = counts.get(key, 0) + 1

    sheet.append(["Placeholder Inventory"])
    sheet["A1"].font = Font(bold=True, size=14)
    sheet.append([])
    sheet.append(["Template", template_name])
    sheet.append(["Unique placeholders", len(placeholders)])
    sheet.append(["Total occurrences", sum(p.count for p in placeholders)])
    sheet.append([])
    sheet.append(["Type", "Count"])
    sheet["A7"].font = _HEADER_FONT
    sheet["B7"].font = _HEADER_FONT
    for key in sorted(counts):
        sheet.append([key, counts[key]])

    sheet.append([])
    sheet.append(["Expected Value and AI Replaced Text are intentionally blank - QA fills them in."])

    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 60


def _write_qa_sheet(workbook: Workbook, rows: list[dict]) -> None:
    sheet = workbook.create_sheet("QA")

    # Row 1 is the group-header band in the real files; the parser skips it.
    sheet.append([])
    sheet.append(HEADERS)
    for cell in sheet[2]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    for row in rows:
        sheet.append(_as_cells(row))

    _finish(sheet)


def _write_type_sheet(workbook: Workbook, name: str, rows: list[dict]) -> None:
    sheet = workbook.create_sheet(name)
    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    for row in rows:
        sheet.append(_as_cells(row))

    _finish(sheet)


def _as_cells(row: dict) -> list:
    return [
        row["name"],        # 0  Placeholder
        "",                 # 1  Column1
        row["type"],        # 2  Placeholder Type
        "",                 # 3  Expected Value      - QA fills
        "",                 # 4  Source Document     - QA fills
        row["instruction"], # 5  Writing instruction - the surrounding sentence
        "",                 # 6  AI Replaced Text    - QA fills
        "",                 # 7  Ai detected Source
        "",                 # 8  Matching Accuracy
        "",                 # 9  Similarity Score
        "",                 # 10 Source Match
    ]


def _finish(sheet) -> None:
    widths = {"A": 52, "B": 8, "C": 16, "D": 30, "E": 24, "F": 70, "G": 30}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    sheet.freeze_panes = sheet["A3"] if sheet.title == "QA" else sheet["A2"]

    for row in sheet.iter_rows(min_row=1):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write(placeholders: list[Placeholder], template_name: str, output_path: str) -> None:
    build(placeholders, template_name).save(output_path)


# ── Reference draft ───────────────────────────────────────────────────────
# A reference file is what QA authors: "this is what each placeholder SHOULD
# contain". Extracting a template gives you every placeholder with the expected
# text still blank - the starting point for that job, not an AgileWriter output.
#
# reference-file-loader.ts reads sheet 0, columns:
#   0 name | 1 type | 2 expected text | 3 source document | 4 notes

REFERENCE_HEADERS = [
    "Placeholder Name",
    "Placeholder Type",
    "Expected Text",
    "Source Document",
    "Notes",
]


def build_reference(placeholders: list[Placeholder]) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "QA"

    sheet.append(REFERENCE_HEADERS)
    for cell in sheet[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    for placeholder in placeholders:
        sheet.append([
            f"<{placeholder.name}>",
            classify(placeholder.name, placeholder.structure),
            "",                      # Expected Text - QA fills this in
            "",                      # Source Document - QA fills this in
            placeholder.context,     # the sentence it sits in, as a hint
        ])

    for column, width in {"A": 52, "B": 16, "C": 60, "D": 26, "E": 70}.items():
        sheet.column_dimensions[column].width = width

    sheet.freeze_panes = sheet["A2"]
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    return workbook


def write_reference(placeholders: list[Placeholder], output_path: str) -> None:
    build_reference(placeholders).save(output_path)
