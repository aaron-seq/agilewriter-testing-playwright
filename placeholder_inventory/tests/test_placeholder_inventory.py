import sys
import zipfile
from pathlib import Path

import pytest

PROJECT_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(PROJECT_ROOT))

from placeholder_inventory.classify import KEY_VALUE, LIST, PARAGRAPH, TABLE, classify
from placeholder_inventory.extractor import extract
from placeholder_inventory.workbook import HEADERS, build

FIXTURES = Path(__file__).resolve().parent.parent / "fixtures"
ICF_TEMPLATE = FIXTURES / "ICF_SET0.docx"
REFERENCE = PROJECT_ROOT / "reference_files" / "ref_ICF_Full.xlsx"

DOC_XML = "word/document.xml"
NS = (
    'xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"'
)


def _docx(tmp_path: Path, body: str, name: str = "t.docx") -> str:
    """Minimal .docx containing the given <w:body> content."""
    path = tmp_path / name
    document = f'<?xml version="1.0"?><w:document {NS}><w:body>{body}</w:body></w:document>'
    with zipfile.ZipFile(path, "w") as archive:
        archive.writestr(DOC_XML, document)
    return str(path)


def _esc(text: str) -> str:
    """Word stores placeholder brackets escaped, as &lt;Sponsor&gt;."""
    return text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _para(*runs: str, props: str = "") -> str:
    body = "".join(f"<w:r><w:t>{_esc(r)}</w:t></w:r>" for r in runs)
    return f"<w:p>{props}{body}</w:p>"


# ── Extraction ────────────────────────────────────────────────────────────

def test_finds_a_simple_placeholder(tmp_path):
    path = _docx(tmp_path, _para("<Sponsor>"))
    assert [p.name for p in extract(path)] == ["Sponsor"]


def test_finds_a_placeholder_split_across_runs(tmp_path):
    """Word splits words across runs constantly; per-run matching finds nothing."""
    path = _docx(tmp_path, _para("<Study ", "Ti", "tle>"))
    assert [p.name for p in extract(path)] == ["Study Title"]


def test_finds_a_placeholder_inside_a_tracked_deletion(tmp_path):
    """Revised templates keep the placeholder in w:delText, not w:t."""
    body = (
        "<w:p><w:del><w:r><w:delText>&lt;Deleted Field&gt;</w:delText></w:r></w:del></w:p>"
    )
    path = _docx(tmp_path, body)
    names = [p.name for p in extract(path)]
    assert "Deleted Field" in names


def test_counts_repeat_occurrences(tmp_path):
    path = _docx(tmp_path, _para("<Sponsor>") + _para("<Sponsor>") + _para("<Other Field>"))
    by_name = {p.name: p for p in extract(path)}
    assert by_name["Sponsor"].count == 2
    assert by_name["Other Field"].count == 1


def test_rejects_noise(tmp_path):
    path = _docx(tmp_path, _para("<insert>", "<711>", "<a>", "<https://x.com>"))
    assert extract(path) == []


def test_reads_headers_and_footers(tmp_path):
    """A scanner limited to document.xml misses everything in the header."""
    path = tmp_path / "h.docx"
    doc = f'<?xml version="1.0"?><w:document {NS}><w:body>{_para("<Body Field>")}</w:body></w:document>'
    hdr = f'<?xml version="1.0"?><w:hdr {NS}>{_para("<Header Field>")}</w:hdr>'
    with zipfile.ZipFile(path, "w") as archive:
        archive.writestr(DOC_XML, doc)
        archive.writestr("word/header1.xml", hdr)

    names = [p.name for p in extract(str(path))]
    assert "Body Field" in names
    assert "Header Field" in names


def test_detects_list_structure(tmp_path):
    numbered = "<w:pPr><w:numPr><w:ilvl w:val='0'/></w:numPr></w:pPr>"
    path = _docx(tmp_path, _para("<Bulleted Item>", props=numbered))
    assert extract(path)[0].structure == "List"


def test_detects_table_structure(tmp_path):
    body = f"<w:tbl><w:tr><w:tc>{_para('<Cell Field>')}</w:tc></w:tr></w:tbl>"
    path = _docx(tmp_path, body)
    assert extract(path)[0].structure == "Table"


def test_survives_a_malformed_part(tmp_path):
    """One broken part must not lose the rest of the document."""
    path = tmp_path / "b.docx"
    doc = f'<?xml version="1.0"?><w:document {NS}><w:body>{_para("<Good Field>")}</w:body></w:document>'
    with zipfile.ZipFile(path, "w") as archive:
        archive.writestr(DOC_XML, doc)
        archive.writestr("word/header1.xml", "<not valid xml")

    assert [p.name for p in extract(str(path))] == ["Good Field"]


# ── Classification ────────────────────────────────────────────────────────

@pytest.mark.parametrize("name,expected", [
    ("Sponsor", KEY_VALUE),
    ("Protocol Number", KEY_VALUE),
    ("Bulleted list of exclusion criteria", LIST),
    ("summarize the endpoints in bullets", LIST),
    ("Table 3: Demographics", TABLE),
    ("brief lay description of disease", PARAGRAPH),
])
def test_classification(name, expected):
    assert classify(name) == expected


def test_the_name_beats_the_structure():
    """
    ICF_SET0 numbers ordinary sections, so <w:numPr> marks most of the body.
    Letting structure win classified nearly everything as List (50% correct);
    letting the name win scores 70%.
    """
    assert classify("Table 3: Demographics", structure=LIST) == TABLE
    assert classify("Sponsor", structure=LIST) == KEY_VALUE


def test_structure_decides_only_when_the_name_is_silent():
    unrevealing = "cohort assignment for the participants enrolled at site"
    assert classify(unrevealing, structure=LIST) == LIST
    assert classify(unrevealing, structure=PARAGRAPH) == PARAGRAPH


# ── Workbook ──────────────────────────────────────────────────────────────

def test_workbook_matches_the_qa_shape(tmp_path):
    path = _docx(tmp_path, _para("<Sponsor>") + _para("<bulleted list of criteria>"))
    workbook = build(extract(path), "t.docx")

    assert "QA" in workbook.sheetnames
    qa = workbook["QA"]

    # accuracy-scorer.ts reads the header from row 2 and data from row 3.
    assert [c.value for c in qa[2]] == HEADERS
    assert qa.cell(row=3, column=1).value == "<Sponsor>"
    assert qa.cell(row=3, column=3).value == KEY_VALUE

    # Values are QA's job, so these stay empty.
    assert qa.cell(row=3, column=4).value == ""   # Expected Value
    assert qa.cell(row=3, column=7).value == ""   # AI Replaced Text


# ── Against the real template ─────────────────────────────────────────────

CSR_TEMPLATE = FIXTURES / "CSR_Template_20FEB2026.docx"
CSR_REFERENCE = PROJECT_ROOT / "reference_files" / "ref_CSR_v3.xlsx"


def _scorer_normalise(value) -> str:
    """The contract normalizePlaceholderName() uses in accuracy-scorer.ts."""
    import re
    text = str(value or "").replace("<", "").replace(">", "").replace("’", "'")
    return re.sub(r"\s+", " ", text).strip().lower()


@pytest.mark.skipif(
    not (CSR_TEMPLATE.is_file() and CSR_REFERENCE.is_file()),
    reason="CSR template or reference not present",
)
def test_extracts_every_reference_placeholder_from_the_csr_template():
    """
    A different template family from ICF, to prove the extractor is not tuned
    to one document. ref_CSR_v3 stores some names with doubled spaces; the
    scorer collapses whitespace before comparing, so this does too.
    """
    import openpyxl

    book = openpyxl.load_workbook(CSR_REFERENCE)
    sheet = book[book.sheetnames[0]]
    expected = {
        _scorer_normalise(row[0])
        for row in sheet.iter_rows(min_row=2, values_only=True)
        if row[0]
    }
    found = {_scorer_normalise(p.name) for p in extract(str(CSR_TEMPLATE))}

    assert expected - found == set(), "placeholders in the reference but not extracted"


@pytest.mark.skipif(not ICF_TEMPLATE.is_file(), reason="ICF template not present")
def test_extracts_every_reference_placeholder_from_the_real_template():
    """
    Every placeholder in ref_ICF_Full.xlsx must be found in ICF_SET0, with one
    known exception: the reference carries a corrupted row reading
    "screeningthe clinic" where the document says "time in the clinic". That
    text appears nowhere in the template, so it cannot be extracted.
    """
    import openpyxl

    normalise = lambda s: str(s or "").replace("<", "").replace(">", "").strip().lower()

    sheet = openpyxl.load_workbook(REFERENCE)[openpyxl.load_workbook(REFERENCE).sheetnames[0]]
    expected = {
        normalise(row[0]) for row in sheet.iter_rows(min_row=2, values_only=True) if row[0]
    }
    KNOWN_BAD_REFERENCE_ROW = (
        "lay summary of duration on study including details/time for screening, "
        "screeningthe clinic, follow-up periods"
    )
    expected.discard(KNOWN_BAD_REFERENCE_ROW)

    found = {normalise(p.name) for p in extract(str(ICF_TEMPLATE))}

    assert expected - found == set(), "placeholders in the reference but not extracted"
